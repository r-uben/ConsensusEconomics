from consensus_economics.utils.date_format import DateFormatUtils
from consensus_economics.paths import Paths



from openpyxl import load_workbook
import pandas as pd


class Worksheet():

    def __init__(self, date: str, country: str) -> None:
        self.__date = date
        self.__year = None
        self.__month = None
        self.__country = country
        self.__workbook = None
        self.__sheets = None
        self.__country_sheet = None
        self.__forecasters_data = None
        self.__variables = None
        self.__column_names = None
        self.__release_date = None

        self.__DATE_FORMAT_CHECKED = False
        self.__FORECASTED_DATA_CLEANED = False

    @property
    def country(self):
        return self.__country

    @property
    def date(self):
        if not self.__DATE_FORMAT_CHECKED:
            if not type(self.__date) == str:
                raise ValueError("Date must be a string")
            if not len(self.__date) == 6:
                raise ValueError("Date must be a 6-digit string: yyyymmdd.")
            self.__DATE_FORMAT_CHECKED = True
        return self.__date
    
    @property
    def DATE_FORMAT_CHECKED(self):
        return self.__DATE_FORMAT_CHECKED
    
    @property
    def FORECASTED_DATA_CLEANED(self):
        return self.__FORECASTED_DATA_CLEANED
    
    @property
    def year(self):
        if self.__year is None:
            self.__year = int(self.date[:4])
        return self.__year
    
    @property
    def month(self):
        if self.__month is None:
            self.__month = int(self.date[4:])
        return self.__month
    
    @property
    def workbook(self):
        if self.__workbook is None: self.get_workbook()
        return self.__workbook
    
    @property
    def sheets(self):
        if self.__sheets is None: self.get_sheets()
        return self.__sheets
    
    @property
    def country_sheet(self):
        if self.__country_sheet is None: self.get_country_sheet()
        return self.__country_sheet
    
    @property
    def forecasters_data(self):
        if self.__forecasters_data is None: self.get_forecasters_data()
        if not self.FORECASTED_DATA_CLEANED: 
            self.clean_forecasters_dataframe()
            self.__FORECASTED_DATA_CLEANED = True
        return self.__forecasters_data
    
    @property
    def column_names(self):
        if self.__column_names is None: self.get_column_names()
        return self.__column_names
    
    @property
    def release_date(self):
        if self.__release_date is None: self.formatted_release_date()
        return self.__release_date
    

    def get_workbook(self) ->  None:
        master_book = load_workbook(f"{Paths().processed}/xlsx/{self.date}.xlsx", data_only=True)
        self.__workbook = master_book

    def get_sheets(self) -> None:
        sheet_names = self.workbook.sheetnames
        sheets = pd.DataFrame(sheet_names, columns=["Sheet Names"])
        self.__sheets = sheets

    def get_country_sheet(self) -> None:
        country_sheet = self.workbook[self.country]
        country_sheet = pd.DataFrame(country_sheet.values)
        self.__country_sheet = country_sheet
    
    def get_column_names(self) -> list:
        # Extract the relevant column names which are spread over multiple rows
        column_headers = self.country_sheet.iloc[1:4]
        # Combine the rows to create a single header row
        column_names = column_headers.apply(lambda x: ' '.join(x.dropna().astype(str).str.strip()), axis=0)
        # Display the new column names
        self.__column_names = column_names.tolist()

    def get_forecasters_data(self) -> None:
        self.country_sheet.columns = self.column_names
        # Extract rows 3-6 for common information
        self.__variables = self.country_sheet.iloc[2:6].reset_index(drop=True)

        # Extract rows 8-13 for mean, high, low, standard deviation, and number of forecasters
        summary_stats = self.country_sheet.iloc[7:13].reset_index(drop=True)

        # Extract rows 26-52 for individual forecasters' values
        forecasters_data = self.country_sheet.iloc[25:].reset_index(drop=True)

        forecasters_data = forecasters_data.iloc[:, [0] + list(range(2, forecasters_data.shape[1]))]
        forecasters_data.columns = [x if x != '' else forecasters_data.columns[i-1] for i, x in enumerate(forecasters_data.columns)]
        forecasters_data.dropna(axis=1, how='all', inplace=True)

        # Add a new row with "year" + [year, year+1]*(number_columns -1)
        new_row = ["year"] + [int(self.year), int(self.year + 1)] * ((forecasters_data.shape[1] - 1) // 2)
        if len(new_row) < forecasters_data.shape[1]:
            new_row += [int(self.year)]  # In case the number of columns is odd, add one more year
        forecasters_data.loc[-1] = new_row  # Add the new row at the top
        forecasters_data.index = forecasters_data.index + 1  # Shift the index
        forecasters_data = forecasters_data.sort_index()  # Sort by index to reset the order
        self.__forecasters_data = forecasters_data


    def clean_forecasters_dataframe(self):
        # Store units from variables DataFrame (rows 2-6)
        units = self.__variables.iloc[2]  # Row with units like "real, % change"
        
        release_date = DateFormatUtils.formatted_release_date(self.__forecasters_data)
        self.__forecasters_data = self.__forecasters_data.set_index(self.__forecasters_data.columns[0])
        self.__forecasters_data.index.name = None
        self.__forecasters_data = self.__forecasters_data.T
        self.__forecasters_data.index.name = "variable"
        self.__forecasters_data = self.__forecasters_data.reset_index()
        self.__forecasters_data = self.__forecasters_data.melt(id_vars=["variable", "year"], var_name="forecaster", value_name="value")
        
        # Add units column by mapping variable names to their corresponding units
        variable_to_unit = dict(zip(units.index, units.values))
        self.__forecasters_data["unit"] = self.__forecasters_data["variable"].map(variable_to_unit)
        
        self.__forecasters_data["year"] = self.__forecasters_data["year"].astype(int)
        self.__forecasters_data["release_date"] = release_date
        self.__forecasters_data["month"] = self.month
        self.__forecasters_data["country"] = self.country
