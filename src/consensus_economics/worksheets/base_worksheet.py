"""Base worksheet class for Consensus Economics Excel files."""

from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from pandas import DataFrame

from consensus_economics.paths import Paths


# Module-level workbook cache to avoid reloading the same file
_workbook_cache: Dict[str, Workbook] = {}


def get_cached_workbook(date: str) -> Workbook:
    """Get a workbook from cache or load it."""
    if date not in _workbook_cache:
        filepath = Paths().xlsx / f"{date}.xlsx"
        _workbook_cache[date] = load_workbook(filepath, data_only=True)
    return _workbook_cache[date]


def clear_workbook_cache(date: Optional[str] = None) -> None:
    """Clear workbook cache. If date provided, clear only that entry."""
    if date:
        _workbook_cache.pop(date, None)
    else:
        _workbook_cache.clear()


class BaseWorksheet:
    """
    Base class for handling consensus economics worksheet operations.

    Args:
        date: Date in format 'yyyymm'
        sheet_name: Sheet name in the workbook
        workbook: Optional pre-loaded workbook (for batch processing)
    """

    def __init__(
        self, date: str, sheet_name: str, workbook: Optional[Workbook] = None
    ) -> None:
        if not isinstance(date, str):
            raise ValueError("Date must be a string")
        if len(date) != 6:
            raise ValueError("Date must be a 6-digit string: yyyymm.")

        if not isinstance(sheet_name, str) or not sheet_name.strip():
            raise ValueError("Sheet name must be a non-empty string")

        self._date = date
        self._year = int(date[:4])
        self._month = int(date[4:])
        self._sheet_name = sheet_name.strip()

        # Use provided workbook or get from cache
        self._workbook: Optional[Workbook] = workbook
        self._sheets: Optional[DataFrame] = None
        self._worksheet: Optional[DataFrame] = None
        self._column_names: Optional[List[str]] = None

    @property
    def sheet_name(self) -> str:
        """The sheet name in the workbook."""
        return self._sheet_name

    @property
    def date(self) -> str:
        """The date string in yyyymm format."""
        return self._date

    @property
    def year(self) -> int:
        """The year component of the date."""
        return self._year

    @property
    def month(self) -> int:
        """The month component of the date."""
        return self._month

    @property
    def workbook(self) -> Workbook:
        """The loaded Excel workbook."""
        if self._workbook is None:
            self._workbook = get_cached_workbook(self._date)
        return self._workbook

    @property
    def sheets(self) -> DataFrame:
        """DataFrame containing sheet names."""
        if self._sheets is None:
            self._sheets = self._get_sheets()
        return self._sheets

    @property
    def worksheet(self) -> DataFrame:
        """DataFrame containing the worksheet's data."""
        if self._worksheet is None:
            self._worksheet = self._get_worksheet()
        return self._worksheet

    def _get_sheets(self) -> DataFrame:
        """Return DataFrame of sheet names."""
        sheet_names = self.workbook.sheetnames
        return DataFrame(sheet_names, columns=["Sheet Names"])

    def _get_worksheet(self) -> DataFrame:
        """Return DataFrame of the worksheet's data."""
        worksheet = self.workbook[self.sheet_name]
        return DataFrame(worksheet.values)
